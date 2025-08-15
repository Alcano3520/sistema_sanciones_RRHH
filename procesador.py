# procesador.py - Lógica de Procesamiento RRHH OPTIMIZADA
import sqlite3
import requests
import json
import threading
import time
from datetime import datetime, date
from typing import List, Dict, Optional, Tuple, Callable
from concurrent.futures import ThreadPoolExecutor, as_completed
from config import *

class ProcesadorRRHH:
    def __init__(self):
        self.supabase_headers = {
            'apikey': SUPABASE_KEY,
            'Authorization': f'Bearer {SUPABASE_KEY}',
            'Content-Type': 'application/json',
            'Prefer': 'return=representation'
        }
        self.init_db_local()
        self._lock = threading.Lock()
        
    def init_db_local(self):
        """Inicializar base de datos local SQLite OPTIMIZADA CON MIGRACIÓN"""
        try:
            with sqlite3.connect(DB_LOCAL) as conn:
                cursor = conn.cursor()
                
                # ⭐ MIGRACIÓN: Verificar y crear/actualizar tablas existentes
                self._migrar_base_datos(cursor)
                
                # Insertar usuarios por defecto
                for user, pwd in USUARIOS_DEFAULT.items():
                    cursor.execute('''
                        INSERT OR IGNORE INTO usuarios (usuario, password) 
                        VALUES (?, ?)
                    ''', (user, pwd))
                
                conn.commit()
                print("✅ Base de datos local inicializada/migrada con optimizaciones")
                
        except Exception as e:
            print(f"❌ Error inicializando DB local: {e}")
            raise
    
    def _migrar_base_datos(self, cursor):
        """🔄 NUEVA: Migrar base de datos existente a nueva estructura"""
        try:
            # 1. Crear tabla procesadas con estructura original
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS procesadas (
                    id TEXT PRIMARY KEY,
                    fecha_proceso DATE DEFAULT CURRENT_TIMESTAMP,
                    usuario TEXT,
                    empleado_cod INTEGER,
                    empleado_nombre TEXT,
                    tipo_sancion TEXT
                )
            ''')
            
            # 2. Agregar nuevas columnas si no existen
            columnas_nuevas = [
                ('fecha_original', 'DATE'),
                ('tiempo_procesamiento', 'REAL')
            ]
            
            for columna, tipo in columnas_nuevas:
                try:
                    cursor.execute(f'ALTER TABLE procesadas ADD COLUMN {columna} {tipo}')
                    print(f"✅ Agregada columna {columna} a tabla procesadas")
                except sqlite3.OperationalError as e:
                    if "duplicate column name" in str(e).lower():
                        pass  # Columna ya existe
                    else:
                        print(f"⚠️ Info columna {columna}: {e}")
            
            # 3. Crear tabla usuarios con estructura básica
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS usuarios (
                    usuario TEXT PRIMARY KEY,
                    password TEXT,
                    activo INTEGER DEFAULT 1
                )
            ''')
            
            # 4. Agregar columna ultimo_acceso si no existe
            try:
                cursor.execute('ALTER TABLE usuarios ADD COLUMN ultimo_acceso DATETIME')
                print("✅ Agregada columna ultimo_acceso a tabla usuarios")
            except sqlite3.OperationalError as e:
                if "duplicate column name" in str(e).lower():
                    pass  # Columna ya existe
                else:
                    print(f"⚠️ Info columna ultimo_acceso: {e}")
            
            # 5. Crear tabla log_operaciones si no existe
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS log_operaciones (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    fecha DATETIME DEFAULT CURRENT_TIMESTAMP,
                    usuario TEXT,
                    operacion TEXT,
                    detalle TEXT,
                    resultado TEXT
                )
            ''')
            
            # 6. Crear índices si no existen (ignora errores si ya existen)
            indices = [
                'CREATE INDEX IF NOT EXISTS idx_procesadas_fecha ON procesadas(fecha_proceso)',
                'CREATE INDEX IF NOT EXISTS idx_procesadas_usuario ON procesadas(usuario)',
                'CREATE INDEX IF NOT EXISTS idx_log_fecha ON log_operaciones(fecha)'
            ]
            
            for indice in indices:
                try:
                    cursor.execute(indice)
                except sqlite3.OperationalError:
                    pass  # Índice ya existe
            
            print("🔄 Migración de base de datos completada")
            
        except Exception as e:
            print(f"⚠️ Error en migración (continuando con estructura básica): {e}")
            # Crear estructura mínima funcional
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS procesadas (
                    id TEXT PRIMARY KEY,
                    fecha_proceso DATE DEFAULT CURRENT_TIMESTAMP,
                    usuario TEXT,
                    empleado_cod INTEGER,
                    empleado_nombre TEXT,
                    tipo_sancion TEXT
                )
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS usuarios (
                    usuario TEXT PRIMARY KEY,
                    password TEXT,
                    activo INTEGER DEFAULT 1
                )
            ''')
    
    def log_operacion(self, usuario: str, operacion: str, detalle: str, resultado: str):
        """Registrar operación en log local COMPATIBLE"""
        try:
            with sqlite3.connect(DB_LOCAL) as conn:
                cursor = conn.cursor()
                
                # 🔄 COMPATIBLE: Verificar si la tabla log_operaciones existe
                cursor.execute('''
                    SELECT name FROM sqlite_master 
                    WHERE type='table' AND name='log_operaciones'
                ''')
                
                if cursor.fetchone():
                    # Tabla existe, registrar log
                    cursor.execute('''
                        INSERT INTO log_operaciones (usuario, operacion, detalle, resultado)
                        VALUES (?, ?, ?, ?)
                    ''', (usuario, operacion, detalle, resultado))
                    conn.commit()
                else:
                    # Tabla no existe, solo imprimir en consola
                    print(f"📝 LOG: {usuario} - {operacion} - {resultado}")
                    
        except Exception as e:
            print(f"⚠️ Error logging: {e}")
            # Fallback: imprimir en consola
            print(f"📝 LOG: {usuario} - {operacion} - {resultado}")
    
    def validar_usuario(self, usuario: str, password: str) -> bool:
        """Validar credenciales de usuario CON COMPATIBILIDAD"""
        try:
            with sqlite3.connect(DB_LOCAL) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT COUNT(*) FROM usuarios 
                    WHERE usuario = ? AND password = ? AND activo = 1
                ''', (usuario, password))
                
                valido = cursor.fetchone()[0] > 0
                
                if valido:
                    # 🔄 COMPATIBLE: Intentar actualizar último acceso si la columna existe
                    try:
                        cursor.execute('''
                            UPDATE usuarios SET ultimo_acceso = CURRENT_TIMESTAMP
                            WHERE usuario = ?
                        ''', (usuario,))
                        conn.commit()
                    except sqlite3.OperationalError:
                        # Columna ultimo_acceso no existe - continuar sin actualizar
                        pass
                
                return valido
                
        except Exception as e:
            print(f"❌ Error validando usuario: {e}")
            return False
    
    def test_conexion_supabase(self) -> bool:
        """Probar conexión con Supabase con timeout"""
        try:
            url = f"{SUPABASE_URL}/rest/v1/sanciones"
            params = {'select': 'count', 'limit': 1}
            
            response = requests.get(
                url, 
                headers=self.supabase_headers, 
                params=params,
                timeout=REQUEST_TIMEOUT
            )
            return response.status_code == 200
            
        except Exception as e:
            print(f"❌ Error conexión Supabase: {e}")
            return False
    
    def obtener_sanciones_pendientes(self) -> List[Dict]:
        """
        ⚡ OPTIMIZADO: Solo consulta Supabase (eliminado doble control)
        Obtener sanciones aprobadas sin comentario RRHH
        """
        try:
            print("🔍 Consultando sanciones pendientes en Supabase...")
            
            url = f"{SUPABASE_URL}/rest/v1/sanciones"
            
            # ⭐ OPTIMIZACIÓN: Query directa sin doble verificación
            response = requests.get(
                url, 
                headers=self.supabase_headers, 
                params=QUERY_PENDIENTES,
                timeout=REQUEST_TIMEOUT
            )
            
            print(f"📡 Status respuesta: {response.status_code}")
            
            if response.status_code != 200:
                print(f"❌ Error API Supabase: {response.status_code}")
                print(f"❌ Respuesta: {response.text}")
                return []
            
            sanciones = response.json()
            print(f"📊 Sanciones pendientes encontradas: {len(sanciones)}")
            
            if len(sanciones) == 0:
                print("ℹ️ No hay sanciones pendientes en este momento")
            
            return sanciones
            
        except requests.Timeout:
            print(f"⏱️ Timeout consultando Supabase ({REQUEST_TIMEOUT}s)")
            return []
        except Exception as e:
            print(f"❌ Error obteniendo sanciones: {e}")
            return []
    
    def validar_disponibilidad_sanciones(self, ids_sanciones: List[str]) -> Tuple[List[str], List[str]]:
        """
        🔒 NUEVO: Validar que las sanciones siguen disponibles (control de concurrencia)
        Retorna: (disponibles, no_disponibles)
        """
        try:
            if not ids_sanciones:
                return [], []
            
            print(f"🔍 Validando disponibilidad de {len(ids_sanciones)} sanciones...")
            
            # Consultar estado actual en Supabase
            url = f"{SUPABASE_URL}/rest/v1/sanciones"
            params = {
                'select': 'id,comentarios_rrhh,updated_at',
                'id': f'in.({",".join(ids_sanciones)})'
            }
            
            response = requests.get(
                url, 
                headers=self.supabase_headers, 
                params=params,
                timeout=REQUEST_TIMEOUT
            )
            
            if response.status_code != 200:
                print(f"❌ Error validando disponibilidad: {response.status_code}")
                return ids_sanciones, []  # Asumir disponibles si falla la validación
            
            sanciones_actuales = response.json()
            
            disponibles = []
            no_disponibles = []
            
            for id_sancion in ids_sanciones:
                sancion_actual = next(
                    (s for s in sanciones_actuales if s['id'] == id_sancion), 
                    None
                )
                
                if sancion_actual and not sancion_actual.get('comentarios_rrhh'):
                    disponibles.append(id_sancion)
                else:
                    no_disponibles.append(id_sancion)
            
            if no_disponibles:
                print(f"⚠️ {len(no_disponibles)} sanciones ya fueron procesadas por otro usuario")
            
            print(f"✅ {len(disponibles)} sanciones disponibles para procesar")
            
            return disponibles, no_disponibles
            
        except Exception as e:
            print(f"❌ Error validando disponibilidad: {e}")
            return ids_sanciones, []  # En caso de error, asumir disponibles
    
    def procesar_sancion_individual(self, sancion: Dict, usuario: str) -> Tuple[bool, str]:
        """
        ⚡ OPTIMIZADO: Procesar una sanción individual con mejor manejo de errores
        Retorna: (éxito, mensaje)
        """
        sancion_id = sancion['id']
        inicio_tiempo = time.time()
        
        try:
            fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M")
            comentario = f"{MSG_PROCESADO} - {fecha_actual} - {usuario}"
            
            print(f"🔄 Procesando {sancion_id[:8]}...")
            
            # Actualizar en Supabase
            url = f"{SUPABASE_URL}/rest/v1/sanciones"
            params = {'id': f'eq.{sancion_id}'}
            data = {'comentarios_rrhh': comentario}
            
            response = requests.patch(
                url, 
                headers=self.supabase_headers, 
                params=params, 
                json=data,
                timeout=REQUEST_TIMEOUT
            )
            
            if response.status_code not in [200, 204]:
                error_msg = f"Error Supabase: {response.status_code}"
                print(f"❌ {error_msg}")
                return False, error_msg
            
            # Guardar en local para historial (sin afectar lógica principal)
            tiempo_procesamiento = time.time() - inicio_tiempo
            self._guardar_procesada_local(sancion, usuario, tiempo_procesamiento)
            
            print(f"✅ Procesada {sancion_id[:8]} en {tiempo_procesamiento:.2f}s")
            return True, "Éxito"
            
        except requests.Timeout:
            error_msg = f"Timeout ({REQUEST_TIMEOUT}s)"
            print(f"⏱️ {error_msg}")
            return False, error_msg
        except Exception as e:
            error_msg = f"Error: {str(e)}"
            print(f"❌ Error procesando {sancion_id[:8]}: {e}")
            return False, error_msg
    
    def _guardar_procesada_local(self, sancion: Dict, usuario: str, tiempo_procesamiento: float):
        """Guardar en base local para historial COMPATIBLE con DB existente"""
        try:
            with sqlite3.connect(DB_LOCAL) as conn:
                cursor = conn.cursor()
                
                # 🔄 COMPATIBLE: Verificar qué columnas existen
                cursor.execute("PRAGMA table_info(procesadas)")
                columnas_existentes = [row[1] for row in cursor.fetchall()]
                
                # Insertar con columnas básicas (siempre existen)
                if 'tiempo_procesamiento' in columnas_existentes and 'fecha_original' in columnas_existentes:
                    # Base de datos nueva con columnas adicionales
                    cursor.execute('''
                        INSERT OR REPLACE INTO procesadas (
                            id, usuario, empleado_cod, empleado_nombre, 
                            tipo_sancion, fecha_original, tiempo_procesamiento
                        ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        sancion['id'],
                        usuario,
                        sancion.get('empleado_cod'),
                        sancion.get('empleado_nombre'),
                        sancion.get('tipo_sancion'),
                        sancion.get('fecha'),
                        tiempo_procesamiento
                    ))
                else:
                    # Base de datos original con estructura básica
                    cursor.execute('''
                        INSERT OR REPLACE INTO procesadas (
                            id, usuario, empleado_cod, empleado_nombre, tipo_sancion
                        ) VALUES (?, ?, ?, ?, ?)
                    ''', (
                        sancion['id'],
                        usuario,
                        sancion.get('empleado_cod'),
                        sancion.get('empleado_nombre'),
                        sancion.get('tipo_sancion')
                    ))
                
                conn.commit()
                
        except Exception as e:
            print(f"⚠️ Error guardando local (no crítico): {e}")
    
    def procesar_multiples_sanciones(self, sanciones: List[Dict], usuario: str, 
                                   callback_progreso: Optional[Callable] = None) -> Tuple[int, int, List[str]]:
        """
        ⚡ SÚPER OPTIMIZADO: Procesamiento por lotes con threading y validación de concurrencia
        """
        if not sanciones:
            return 0, 0, []
        
        print(f"🚀 Iniciando procesamiento optimizado de {len(sanciones)} sanciones...")
        inicio_total = time.time()
        
        # 1. VALIDAR DISPONIBILIDAD (prevenir concurrencia)
        ids_sanciones = [s['id'] for s in sanciones]
        ids_disponibles, ids_no_disponibles = self.validar_disponibilidad_sanciones(ids_sanciones)
        
        if callback_progreso:
            callback_progreso(f"Validadas: {len(ids_disponibles)} disponibles, {len(ids_no_disponibles)} ocupadas")
        
        # Filtrar solo sanciones disponibles
        sanciones_disponibles = [s for s in sanciones if s['id'] in ids_disponibles]
        
        if not sanciones_disponibles:
            return 0, len(sanciones), ["Todas las sanciones ya fueron procesadas por otros usuarios"]
        
        exitosas = 0
        fallidas = 0
        errores = []
        
        # 2. PROCESAMIENTO EN LOTES CON THREADING
        def procesar_lote(lote_sanciones):
            lote_exitosas = 0
            lote_fallidas = 0
            lote_errores = []
            
            for sancion in lote_sanciones:
                éxito, mensaje = self.procesar_sancion_individual(sancion, usuario)
                if éxito:
                    lote_exitosas += 1
                else:
                    lote_fallidas += 1
                    lote_errores.append(f"{sancion['id'][:8]}: {mensaje}")
                
                # Callback de progreso
                if callback_progreso:
                    callback_progreso(f"Procesadas: {lote_exitosas + lote_fallidas}/{len(sanciones_disponibles)}")
            
            return lote_exitosas, lote_fallidas, lote_errores
        
        # Dividir en lotes
        lotes = [sanciones_disponibles[i:i + BATCH_SIZE] 
                for i in range(0, len(sanciones_disponibles), BATCH_SIZE)]
        
        print(f"📦 Procesando en {len(lotes)} lotes de máximo {BATCH_SIZE} sanciones")
        
        # Procesar lotes con ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=min(MAX_THREADS, len(lotes))) as executor:
            futures = {executor.submit(procesar_lote, lote): i for i, lote in enumerate(lotes)}
            
            for future in as_completed(futures):
                try:
                    lote_exitosas, lote_fallidas, lote_errores = future.result()
                    exitosas += lote_exitosas
                    fallidas += lote_fallidas
                    errores.extend(lote_errores)
                    
                    if callback_progreso:
                        callback_progreso(f"Lote completado. Total: {exitosas + fallidas}/{len(sanciones_disponibles)}")
                        
                except Exception as e:
                    print(f"❌ Error en lote: {e}")
                    fallidas += BATCH_SIZE
                    errores.append(f"Error en lote: {str(e)}")
        
        tiempo_total = time.time() - inicio_total
        
        # Log de la operación
        self.log_operacion(
            usuario, 
            "PROCESAMIENTO_MASIVO", 
            f"{len(sanciones)} sanciones solicitadas, {len(sanciones_disponibles)} procesadas",
            f"Exitosas: {exitosas}, Fallidas: {fallidas}, Tiempo: {tiempo_total:.2f}s"
        )
        
        print(f"🎯 Procesamiento completado en {tiempo_total:.2f}s")
        print(f"✅ Exitosas: {exitosas}")
        print(f"❌ Fallidas: {fallidas}")
        print(f"⚠️ No disponibles: {len(ids_no_disponibles)}")
        
        return exitosas, fallidas, errores
    
    def categorizar_sanciones(self, sanciones: List[Dict]) -> Dict[str, List[Dict]]:
        """Organizar sanciones por categorías"""
        categorizadas = {categoria: [] for categoria in CATEGORIAS.keys()}
        
        for sancion in sanciones:
            tipo = sancion.get('tipo_sancion', '').upper()
            categorizada = False
            
            for categoria, tipos in CATEGORIAS.items():
                if tipo in tipos:
                    categorizadas[categoria].append(sancion)
                    categorizada = True
                    break
            
            if not categorizada:
                categorizadas["Resto"].append(sancion)
        
        return categorizadas
    
    def obtener_procesadas_completas(self) -> List[Dict]:
        """
        ⚡ OPTIMIZADO: Obtener historial de procesadas desde Supabase
        """
        try:
            print("📚 Consultando historial de sanciones procesadas...")
            
            url = f"{SUPABASE_URL}/rest/v1/sanciones"
            params = {
                'select': '*',
                'comentarios_rrhh': 'not.is.null',
                'order': 'updated_at.desc',
                'limit': 1000  # Limitar para rendimiento
            }
            
            response = requests.get(
                url, 
                headers=self.supabase_headers, 
                params=params,
                timeout=REQUEST_TIMEOUT
            )
            
            if response.status_code == 200:
                procesadas = response.json()
                
                # Enriquecer con datos locales si existen
                self._enriquecer_con_datos_locales(procesadas)
                
                print(f"📊 Historial obtenido: {len(procesadas)} sanciones procesadas")
                return procesadas
            else:
                print(f"❌ Error obteniendo historial: {response.status_code}")
                return []
                
        except Exception as e:
            print(f"❌ Error obteniendo procesadas: {e}")
            return []
    
    def _enriquecer_con_datos_locales(self, sanciones: List[Dict]):
        """Agregar datos locales de procesamiento si existen COMPATIBLE"""
        try:
            with sqlite3.connect(DB_LOCAL) as conn:
                cursor = conn.cursor()
                
                # Verificar qué columnas existen
                cursor.execute("PRAGMA table_info(procesadas)")
                columnas_existentes = [row[1] for row in cursor.fetchall()]
                
                # Query base
                if 'tiempo_procesamiento' in columnas_existentes:
                    cursor.execute('''
                        SELECT id, usuario, fecha_proceso, tiempo_procesamiento
                        FROM procesadas
                    ''')
                    datos_locales = {row[0]: {
                        'procesado_por': row[1], 
                        'fecha_procesamiento': row[2],
                        'tiempo_procesamiento': row[3]
                    } for row in cursor.fetchall()}
                else:
                    # Base de datos original
                    cursor.execute('''
                        SELECT id, usuario, fecha_proceso
                        FROM procesadas
                    ''')
                    datos_locales = {row[0]: {
                        'procesado_por': row[1], 
                        'fecha_procesamiento': row[2],
                        'tiempo_procesamiento': None
                    } for row in cursor.fetchall()}
                
                # Enriquecer sanciones
                for sancion in sanciones:
                    if sancion['id'] in datos_locales:
                        sancion.update(datos_locales[sancion['id']])
                    else:
                        # Extraer usuario del comentario si no hay dato local
                        comentario = sancion.get('comentarios_rrhh', '')
                        if comentario and ' - ' in comentario:
                            partes = comentario.split(' - ')
                            if len(partes) >= 3:
                                sancion['procesado_por'] = partes[-1]
                                sancion['fecha_procesamiento'] = partes[-2]
                
        except Exception as e:
            print(f"⚠️ Error enriqueciendo datos: {e}")
    
    def categorizar_procesadas(self, sanciones: List[Dict]) -> Dict[str, List[Dict]]:
        """Organizar sanciones procesadas por categorías"""
        return self.categorizar_sanciones(sanciones)
    
    def exportar_a_excel(self, sanciones: List[Dict], nombre_archivo: str = None) -> str:
        """⚡ OPTIMIZADO: Exportar sanciones a Excel con mejor rendimiento"""
        try:
            print("📥 Iniciando exportación optimizada a Excel...")
            
            # Verificar e instalar openpyxl si es necesario
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill
                from openpyxl.utils import get_column_letter
            except ImportError:
                print("📦 Instalando openpyxl...")
                import subprocess
                import sys
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill
                from openpyxl.utils import get_column_letter
        
            if not nombre_archivo:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                nombre_archivo = f"Sanciones_Procesadas_{timestamp}.xlsx"
            
            print(f"📄 Creando archivo: {nombre_archivo}")
            
            wb = openpyxl.Workbook()
            categorizadas = self.categorizar_sanciones(sanciones)
            
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Crear hojas por categoría
            hojas_creadas = 0
            for categoria, sanciones_categoria in categorizadas.items():
                if not sanciones_categoria:
                    continue
                    
                print(f"📋 Creando hoja: {categoria} ({len(sanciones_categoria)} registros)")
                
                nombre_hoja = categoria.replace('/', '-')[:30]
                ws = wb.create_sheet(title=nombre_hoja)
                
                # Estilos
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Headers optimizados
                headers = [
                    'ID', 'Empleado Cod', 'Empleado Nombre', 'Puesto', 
                    'Agente', 'Fecha', 'Hora', 'Tipo Sanción', 
                    'Observaciones', 'Status', 'Comentarios RRHH', 
                    'Procesado Por', 'Fecha Procesamiento'
                ]
                
                # Escribir headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Escribir datos en lotes para mejor rendimiento
                batch_data = []
                for sancion in sanciones_categoria:
                    row_data = [
                        str(sancion.get('id', ''))[:12] + '...' if len(str(sancion.get('id', ''))) > 12 else str(sancion.get('id', '')),
                        sancion.get('empleado_cod', ''),
                        sancion.get('empleado_nombre', ''),
                        sancion.get('puesto', ''),
                        sancion.get('agente', ''),
                        sancion.get('fecha', ''),
                        sancion.get('hora', ''),
                        sancion.get('tipo_sancion', ''),
                        sancion.get('observaciones', ''),
                        sancion.get('status', ''),
                        sancion.get('comentarios_rrhh', ''),
                        sancion.get('procesado_por', ''),
                        sancion.get('fecha_procesamiento', '')
                    ]
                    batch_data.append(row_data)
                
                # Escribir todos los datos de una vez
                for row_idx, row_data in enumerate(batch_data, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Ajustar anchos
                column_widths = [15, 12, 25, 20, 15, 12, 10, 18, 30, 12, 30, 15, 18]
                for col, width in enumerate(column_widths, 1):
                    ws.column_dimensions[get_column_letter(col)].width = width
                
                hojas_creadas += 1
            
            # Hoja resumen
            if hojas_creadas > 0:
                ws_resumen = wb.create_sheet(title="Resumen", index=0)
                
                ws_resumen['A1'] = "RESUMEN DE SANCIONES PROCESADAS"
                ws_resumen['A1'].font = Font(size=16, bold=True)
                
                ws_resumen['A3'] = f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws_resumen['A4'] = f"Total de sanciones: {len(sanciones)}"
                
                row = 6
                for categoria, sanciones_cat in categorizadas.items():
                    if sanciones_cat:
                        ws_resumen[f'A{row}'] = f"{categoria}: {len(sanciones_cat)} sanciones"
                        row += 1
            else:
                ws = wb.create_sheet(title="Sin Datos")
                ws['A1'] = "No hay datos para exportar"
            
            print(f"💾 Guardando archivo: {nombre_archivo}")
            wb.save(nombre_archivo)
            print(f"✅ Excel exportado exitosamente")
            
            return nombre_archivo
            
        except Exception as e:
            print(f"❌ Error exportando a Excel: {e}")
            import traceback
            traceback.print_exc()
            return None

    def obtener_estadisticas(self) -> Dict:
        """⚡ OPTIMIZADO: Estadísticas del sistema"""
        try:
            with sqlite3.connect(DB_LOCAL) as conn:
                cursor = conn.cursor()
                
                # Estadísticas locales
                cursor.execute('SELECT COUNT(*) FROM procesadas')
                total_procesadas = cursor.fetchone()[0]
                
                cursor.execute('''
                    SELECT usuario, COUNT(*) 
                    FROM procesadas 
                    GROUP BY usuario
                ''')
                por_usuario = dict(cursor.fetchall())
                
                cursor.execute('''
                    SELECT tipo_sancion, COUNT(*) 
                    FROM procesadas 
                    GROUP BY tipo_sancion
                    ORDER BY COUNT(*) DESC
                ''')
                por_tipo = dict(cursor.fetchall())
                
                cursor.execute('''
                    SELECT COUNT(*) FROM procesadas 
                    WHERE DATE(fecha_proceso) = DATE('now')
                ''')
                hoy = cursor.fetchone()[0]
                
                # Estadísticas de rendimiento
                cursor.execute('''
                    SELECT AVG(tiempo_procesamiento), MIN(tiempo_procesamiento), MAX(tiempo_procesamiento)
                    FROM procesadas 
                    WHERE tiempo_procesamiento IS NOT NULL
                ''')
                tiempos = cursor.fetchone()
                
                return {
                    'total_procesadas': total_procesadas,
                    'por_usuario': por_usuario,
                    'por_tipo': por_tipo,
                    'procesadas_hoy': hoy,
                    'tiempo_promedio': round(tiempos[0] or 0, 2),
                    'tiempo_minimo': round(tiempos[1] or 0, 2),
                    'tiempo_maximo': round(tiempos[2] or 0, 2)
                }
                
        except Exception as e:
            print(f"❌ Error obteniendo estadísticas: {e}")
            return {}

    def crear_sancion_prueba(self) -> bool:
        """Crear sanción de prueba"""
        try:
            print("🧪 Creando sanción de prueba...")
            
            url = f"{SUPABASE_URL}/rest/v1/sanciones"
            
            sancion_prueba = {
                "empleado_cod": 99999,
                "empleado_nombre": "EMPLEADO PRUEBA SISTEMA",
                "puesto": "VIGILANTE DE PRUEBA",
                "agente": "SISTEMA AUTOMATICO",
                "fecha": date.today().isoformat(),
                "hora": "08:00:00",
                "tipo_sancion": "ATRASO",
                "observaciones": f"Sanción de prueba creada el {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                "status": "aprobado",
                "comentarios_rrhh": None,
                "supervisor_id": "00000000-0000-0000-0000-000000000000"
            }
            
            response = requests.post(
                url, 
                headers=self.supabase_headers, 
                json=sancion_prueba,
                timeout=REQUEST_TIMEOUT
            )
            
            if response.status_code in [200, 201]:
                print("✅ Sanción de prueba creada exitosamente")
                return True
            else:
                print(f"❌ Error creando sanción de prueba: {response.status_code}")
                return False
                
        except Exception as e:
            print(f"❌ Error creando sanción de prueba: {e}")
            return False

# Instancia global
procesador = ProcesadorRRHH()
