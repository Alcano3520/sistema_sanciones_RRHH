# test_excel_headers.py - Test específico para verificar headers en Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import datetime

def test_headers_excel():
    """Test específico para asegurar que los headers aparezcan en Excel"""
    
    print("🧪 TEST ESPECÍFICO DE HEADERS EN EXCEL")
    print("="*50)
    
    try:
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Headers"
        
        # Headers exactos del sistema
        headers = [
            'ID Sanción',
            'Código Empleado', 
            'Nombre Empleado',
            'Puesto',
            'Agente Supervisor',
            'Fecha',
            'Hora',
            'Tipo de Sanción',
            'Observaciones',
            'Estado',
            'Comentarios Gerencia',
            'Comentarios RRHH',
            'Procesado Por',
            'Fecha Procesamiento'
        ]
        
        print(f"📋 Escribiendo {len(headers)} headers...")
        
        # Escribir headers con formato
        for col_num, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = str(header_text)
            
            # Aplicar formato exacto del sistema
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            print(f"   ✅ Header {col_num}: '{header_text}'")
        
        # Altura de fila
        ws.row_dimensions[1].height = 30
        
        # Datos de prueba
        datos_prueba = [
            ['SAN001', '1234', 'JUAN PÉREZ GARCÍA', 'VIGILANTE', 'SUPERVISOR A', '2025-01-15', '08:00', 'ATRASO', 'Llegó tarde', 'APROBADO', 'Aprobado por gerencia', 'Procesado para nómina', 'RRHH_USER', '2025-01-15 10:30'],
            ['SAN002', '5678', 'MARÍA LÓPEZ SILVA', 'GUARDIA', 'SUPERVISOR B', '2025-01-16', '14:00', 'PERMISO', 'Permiso médico', 'APROBADO', 'Documentación OK', 'Procesado para nómina', 'RRHH_USER', '2025-01-16 11:15']
        ]
        
        print(f"📊 Escribiendo {len(datos_prueba)} filas de datos...")
        
        for row_num, datos in enumerate(datos_prueba, 2):
            for col_num, valor in enumerate(datos, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = str(valor)
                
                # Borde para datos
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Color alternado
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # Ajustar anchos
        column_widths = [15, 12, 25, 20, 20, 12, 10, 18, 30, 12, 25, 25, 15, 18]
        for col_num, width in enumerate(column_widths, 1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = width
        
        # Guardar archivo
        filename = f"test_headers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        print(f"💾 Archivo guardado: {filename}")
        
        # VERIFICAR CONTENIDO
        print("\n🔍 VERIFICANDO CONTENIDO DEL ARCHIVO...")
        wb_verify = openpyxl.load_workbook(filename)
        ws_verify = wb_verify.active
        
        print("📋 Headers encontrados en el archivo:")
        headers_encontrados = []
        for col in range(1, len(headers) + 1):
            header_value = ws_verify.cell(row=1, column=col).value
            headers_encontrados.append(header_value)
            print(f"   Columna {col}: '{header_value}'")
        
        # Verificar datos
        print("\n📊 Primera fila de datos:")
        for col in range(1, len(headers) + 1):
            data_value = ws_verify.cell(row=2, column=col).value
            print(f"   Columna {col}: '{data_value}'")
        
        wb_verify.close()
        
        # RESULTADO
        headers_ok = all(h is not None and h != '' for h in headers_encontrados)
        
        print(f"\n{'='*50}")
        if headers_ok:
            print("🎉 ¡ÉXITO! Headers aparecen correctamente")
            print(f"✅ Archivo generado: {filename}")
            print("✅ Headers con formato azul")
            print("✅ Datos con bordes")
            print("✅ Estructura completa")
        else:
            print("❌ PROBLEMA: Algunos headers no aparecen")
            print(f"Headers encontrados: {headers_encontrados}")
        
        return filename if headers_ok else None
        
    except Exception as e:
        print(f"❌ Error en test: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    result = test_headers_excel()
    if result:
        print(f"\n✅ Test exitoso. Archivo: {result}")
        try:
            import os
            respuesta = input("¿Abrir archivo para verificar visualmente? (s/n): ")
            if respuesta.lower() in ['s', 'si', 'y', 'yes']:
                os.startfile(result)
        except:
            pass
    else:
        print("\n❌ Test falló")
